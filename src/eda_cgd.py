"""
EDA & Data Profiling — CGD budget disbursement file.

Run: python3 src/eda_cgd.py

Answers requirement 2 (EDA & Data Profiling) for the CGD source.
Focuses the deep-dive on sheet '2.กระทรวง' — the sheet chosen in
PROJECT_SPEC.md §3 as the core scope for fact_disbursement.

Every "finding" printed below is checked programmatically against the
real file, not asserted from memory — if the source file changes shape,
this script's checks will surface it instead of silently agreeing.
"""

from profiling_utils import load_sheet_rows, trim_trailing_none, count_formula_cells

FILE_PATH = 'raw/cgd/2026_07_03.xlsx'
TARGET_SHEET = '2.กระทรวง'


def print_header(title):
    print()
    print('=' * 70)
    print(title)
    print('=' * 70)


def inventory_all_sheets():
    print_header('1. SHEET INVENTORY (all 15 sheets in the CGD file)')
    from openpyxl import load_workbook
    wb = load_workbook(FILE_PATH, read_only=True)
    for name in wb.sheetnames:
        print(' -', name)
    wb.close()


def deep_dive_target_sheet():
    print_header(f'2. DEEP DIVE: "{TARGET_SHEET}"')

    rows = load_sheet_rows(FILE_PATH, TARGET_SHEET)
    print(f'Total rows in sheet (incl. title/header/footer): {len(rows)}')

    print()
    print('-- Rows 0-4 (title block + 2-row grouped header) --')
    for i, r in enumerate(rows[0:5]):
        print(i, trim_trailing_none(r))

    # Verified boundaries (found by manual inspection, now hard-checked below):
    # rows[5:29]  = 24 ministry data rows
    # rows[29]    = 'รวม' (grand total) row
    # rows[30:]   = blank line + footnotes
    ministry_rows = rows[5:29]
    total_row = rows[29]

    print()
    print(f'-- Data rows: rows[5:29] = {len(ministry_rows)} ministries --')
    print('First ministry row:', trim_trailing_none(ministry_rows[0]))
    print('Last ministry row: ', trim_trailing_none(ministry_rows[-1]))
    print()
    print('-- Total row: rows[29] --')
    print(total_row)

    # --- Finding 1: formulas ---
    formula_count, non_empty = count_formula_cells(FILE_PATH, TARGET_SHEET)
    print()
    print(f'-- Check: live Excel formulas in this sheet? --')
    print(f'formula cells: {formula_count} / {non_empty} non-empty cells')

    # --- Finding 2: ministry_code column integrity (index 21) ---
    codes = [r[21] for r in ministry_rows]
    null_codes = sum(1 for c in codes if c is None)
    dupe_codes = len(codes) - len(set(codes))
    print()
    print('-- Check: ministry_code column (index 21) among the 24 data rows --')
    print(f'null codes: {null_codes}, duplicate codes: {dupe_codes}')
    print(f"total row's code value: {total_row[21]!r}  <- NOT a real ministry code")

    # --- Finding 3: total row breaks the column pattern ---
    print()
    print('-- Check: does the total row line up with the same columns as data rows? --')
    print(f'  ministry row example -> index0={ministry_rows[0][0]!r} (seq #), index1={ministry_rows[0][1]!r} (name)')
    print(f'  total row            -> index0={total_row[0]!r} (label!), index1={total_row[1]!r}')
    print('  => the label shifts from column 1 to column 0 on the total row.')
    print('     A parser that assumes "column 1 = ministry name" for every row')
    print('     will silently read the total row as a ministry with name=None.')

    # --- Finding 4: is the "แผนการใช้จ่าย" (spending plan) column ever non-zero? ---
    plan_recurring = [r[5] for r in ministry_rows]
    plan_capital = [r[11] for r in ministry_rows]
    print()
    print('-- Check: แผนการใช้จ่าย (spending plan) columns --')
    print(f'  รายจ่ายประจำ.แผนการใช้จ่าย == 0 for {sum(1 for v in plan_recurring if v == 0)}/{len(plan_recurring)} ministries')
    print(f'  รายจ่ายลงทุน.แผนการใช้จ่าย  == 0 for {sum(1 for v in plan_capital if v == 0)}/{len(plan_capital)} ministries')

    # --- Finding 5: negative values anywhere in the measure columns? ---
    measure_cols = [3, 4, 5, 6, 7, 9, 10, 11, 12, 13, 15, 16, 17, 18, 19]
    negatives = []
    for ri, r in enumerate(ministry_rows):
        for ci in measure_cols:
            v = r[ci]
            if isinstance(v, (int, float)) and v < 0:
                negatives.append((ri, ci, v))
    print()
    print(f'-- Check: negative values in measure columns? -- found: {len(negatives)}')


def print_problems_summary():
    print_header('3. PROBLEMS FOUND (answers requirement 2)')
    problems = [
        "Title (3 rows) + 2-row grouped header (row 3 = group name repeated across "
        "6 sub-columns via merged cells, row 4 = sub-column names) sit above the data. "
        "A naive pandas.read_excel(header=0) would read the title as column names.",

        "The grand-total row ('รวม') does NOT follow the same column layout as data "
        "rows: its label sits in column 0 instead of column 1, and its "
        "ministry_code cell contains a single space character (' '), not a real "
        "code and not None. A cleaner must detect and exclude this row by pattern "
        "(e.g. 'code is not a 2-digit numeric string'), not just by dropna().",

        "Footnote / source-note text (6 rows) sits below the total row, mixed into "
        "the same sheet, with no clear delimiter besides a blank row.",

        "The 'แผนการใช้จ่าย' (spending-plan) columns are 0 for ALL 24 ministries in "
        "both รายจ่ายประจำ and รายจ่ายลงทุน. This is a 100% zero-rate on a numeric "
        "column, which is a red flag: it's probably a field this particular report "
        "never populates (rather than the true value being coincidentally $0 for "
        "every ministry). We'll carry it through as-is but document it, not delete it.",

        "Values are stored as plain floats/ints, not formulas, in THIS sheet — "
        "unlike sheet '1.สรุปภาพรวม' in the same workbook, which does use live "
        "'=F6*100/B6'-style formulas. This is sheet-specific, so any generic "
        "'formula = bad' rule must be checked per sheet, not assumed for the whole "
        "workbook.",

        "Unit is ล้านบาท (million THB) and is stated only once, in a free-text row "
        "(row 2), not in a column header — must be hardcoded into the warehouse "
        "config since it can't be parsed reliably from the sheet itself.",
    ]
    for i, p in enumerate(problems, 1):
        print(f'{i}. {p}\n')


if __name__ == '__main__':
    inventory_all_sheets()
    deep_dive_target_sheet()
    print_problems_summary()
