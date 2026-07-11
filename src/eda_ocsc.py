"""EDA & Data Profiling -- OCSC workforce yearbook. Run:
python3 src/eda_ocsc.py. (1) Classifies all ~68 sheets, tabular vs.
chart-only -- evidence for the PROJECT_SPEC.md §3 scope decision. (2)
Deep-dives sheet '12', checking the flat-looking table is actually a
2-level hierarchy of totals."""

from openpyxl import load_workbook
from profiling_utils import classify_sheet_density, trim_trailing_none

FILE_PATH = 'raw/ocsc/thaigovmanpower2567_4.xlsx'
TARGET_SHEET = '12'

# Threshold used to call a sheet "chart-only / not usable as a table".
# Chosen because every chart-only sheet we found has <= 2 rows with any
# cell value at all (just the title, sometimes a subtitle) — a real data
# table always has a header row + multiple data rows.
CHART_ONLY_THRESHOLD = 2


def print_header(title):
    print()
    print('=' * 70)
    print(title)
    print('=' * 70)


def inventory_all_sheets():
    print_header('1. SHEET INVENTORY (all ~68 sheets in the OCSC file)')

    wb = load_workbook(FILE_PATH, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    results = [classify_sheet_density(FILE_PATH, name) for name in sheet_names]

    usable = [r for r in results if r['rows_with_data'] > CHART_ONLY_THRESHOLD]
    chart_only = [r for r in results if r['rows_with_data'] <= CHART_ONLY_THRESHOLD]

    print(f'Total sheets: {len(results)}')
    print(f'Sheets with rows_with_data > {CHART_ONLY_THRESHOLD} (usable as a table): {len(usable)}')
    print(f'Sheets with rows_with_data <= {CHART_ONLY_THRESHOLD} (chart-only / cover / empty): {len(chart_only)}')

    print()
    print('-- Chart-only / near-empty sheets (excluded from scope, see PROJECT_SPEC.md §3) --')
    for r in chart_only:
        print(f"  {r['sheet']!r:12} rows_with_data={r['rows_with_data']:2}  \"{r['first_cell_text']}\"")

    return results


def deep_dive_target_sheet():
    print_header(f'2. DEEP DIVE: sheet "{TARGET_SHEET}"')

    wb = load_workbook(FILE_PATH, read_only=True)
    ws = wb[TARGET_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    print(f'Total rows in sheet: {len(rows)}')
    print()
    print('-- Full content (trimmed) --')
    for i, r in enumerate(rows):
        vals = trim_trailing_none(r)
        if vals:
            print(i, vals)

    # Verified boundaries (row indices), found by manual inspection above:
    # row 3            = grand total ('รวมทั้งหมด')
    # row 4             = subtotal 'ข้าราชการ'
    # rows 5-18 (14)    = children of ข้าราชการ
    # row 19            = subtotal 'กำลังคนประเภทอื่น'
    # rows 20-27 (8)    = children of กำลังคนประเภทอื่น
    # row 28            = footnote text, not data
    grand_total_row = rows[3]
    subtotal_1_row = rows[4]
    children_1 = rows[5:19]
    subtotal_2_row = rows[19]
    children_2 = rows[20:28]

    print()
    print('-- Check: is this really a flat list, or a hidden hierarchy? --')
    sum_children_1 = sum(r[1] for r in children_1)
    sum_children_2 = sum(r[1] for r in children_2)
    subtotal_1 = subtotal_1_row[1]
    subtotal_2 = subtotal_2_row[1]
    grand_total = grand_total_row[1]

    print(f"  sum of 14 children under '{subtotal_1_row[0]}': {sum_children_1}  "
          f"vs subtotal row value: {subtotal_1}  -> match: {sum_children_1 == subtotal_1}")
    print(f"  sum of 8 children under '{subtotal_2_row[0]}': {sum_children_2}  "
          f"vs subtotal row value: {subtotal_2}  -> match: {sum_children_2 == subtotal_2}")
    print(f"  subtotal_1 + subtotal_2 = {subtotal_1 + subtotal_2}  "
          f"vs grand total row value: {grand_total}  -> match: {subtotal_1 + subtotal_2 == grand_total}")

    if sum_children_1 == subtotal_1 and sum_children_2 == subtotal_2 and subtotal_1 + subtotal_2 == grand_total:
        print()
        print('  CONFIRMED: this is a 2-level hierarchy (grand total -> 2 subtotals -> 22 leaf')
        print('  categories), not 29 independent categories. SUM(headcount) over all rows as-is')
        print('  would overcount by roughly 3x (total + 2 subtotals + their children all add up).')

    # Whitespace check on category name strings
    print()
    print('-- Check: stray whitespace on category names --')
    all_rows = [grand_total_row, subtotal_1_row] + list(children_1) + [subtotal_2_row] + list(children_2)
    dirty = [r[0] for r in all_rows if isinstance(r[0], str) and r[0] != r[0].strip()]
    print(f'  category names with leading/trailing whitespace: {dirty}')


def print_problems_summary():
    print_header('3. PROBLEMS FOUND (answers requirement 2)')
    problems = [
        "Out of ~68 sheets, most are chart-only (title text but no tabular cell "
        "data) or cover/TOC pages. See the inventory above for the exact list. "
        "This is the concrete evidence behind scoping to sheet '12' as core and "
        "treating the rest as out-of-scope (PROJECT_SPEC.md §3) rather than a "
        "guess.",

        "Sheet names are page numbers (e.g. '17-29', '99-113'), not descriptive "
        "labels — you cannot tell what a sheet contains from its name alone. The "
        "'สารบัญ' (table of contents) sheet is the only map from topic to page "
        "number, so any tool selecting sheets by topic must parse it first.",

        "Sheet '12' LOOKS like a flat 3-column table (category / headcount / "
        "percent) but is actually a 2-level hierarchy: 1 grand total row, 2 "
        "subtotal rows, and 22 leaf category rows, all mixed together with no "
        "column indicating hierarchy level. Verified programmatically above: "
        "child rows sum exactly to their subtotal, and both subtotals sum "
        "exactly to the grand total. A naive SUM(headcount) over all 29 rows "
        "overcounts by ~3x. The cleaner must tag each row's hierarchy level (or "
        "keep only leaf rows) before loading — this changes the fact table design "
        "from what PROJECT_SPEC.md §4.2 currently assumes ('flat list').",

        "At least one category name has trailing whitespace "
        "('องค์กรอิสระตามรัฐธรรมนูญ ') — must .strip() every text field before "
        "using it as a dimension key, or it will create a duplicate dimension "
        "row that differs only by invisible whitespace.",

        "A footnote sentence ('หมายเหตุ : ทหาร ประกอบด้วย...') sits in the same "
        "column as category names, one row below the last real data row, with no "
        "structural marker separating it from data — must be excluded by row "
        "position, not by any content rule (it has no keyword that reliably "
        "identifies it as a footnote everywhere in this file).",

        "Percent values are long, unrounded floats (e.g. 58.466126474254324) — "
        "fine to store as-is (more precision is not a problem), but must be "
        "rounded only at presentation time, not in the stored value, or "
        "re-deriving percent-of-total later would compound rounding error.",
    ]
    for i, p in enumerate(problems, 1):
        print(f'{i}. {p}\n')


if __name__ == '__main__':
    inventory_all_sheets()
    deep_dive_target_sheet()
    print_problems_summary()
