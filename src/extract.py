"""Step 3: extractor.

Reads the two source Excel files using per-source config (config/cgd.yaml,
config/ocsc.yaml — sheet name, row boundaries, column indices) and writes
flat CSVs to staging/. This step only gets the right rows/columns out; it
does not clean them (no .strip(), no percent recompute, no total-row math,
no is_leaf). That boundary is deliberate — extraction and cleaning are
scored separately (PROJECT_SPEC.md §1), and Step 4 owns cleaning.

Paths are resolved relative to this file, not the working directory (see
HANDOFF.md's note on the repo-root path bug from Step 1).
"""

from pathlib import Path
import csv

import yaml
from openpyxl import load_workbook

from validate_structure import validate_cgd_structure

REPO_ROOT = Path(__file__).resolve().parent.parent
CONFIG_DIR = REPO_ROOT / "config"
RAW_DIR = REPO_ROOT / "raw"
STAGING_DIR = REPO_ROOT / "staging"


def load_config(source: str) -> dict:
    path = CONFIG_DIR / f"{source}.yaml"
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def find_raw_file(source: str) -> Path:
    """Pick the .xlsx file to extract from raw/<source>/.

    Not hardcoded to today's filename: Step 6's detector will drop new
    files into this same folder under their own report-date name, and this
    extractor should pick those up without a code change.
    """
    source_dir = RAW_DIR / source
    files = sorted(source_dir.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx file found in {source_dir}")
    return files[-1]


def load_sheet_rows(file_path: Path, sheet_name: str) -> list:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def extract_cgd(raw_file: Path, config: dict) -> list[dict]:
    rows = load_sheet_rows(raw_file, config["sheet_name"])
    cols = config["columns"]
    ministry_code_col = config["ministry_code_col"]

    records = []
    for r in rows[config["data_row_start"]: config["data_row_end"] + 1]:
        record = {
            "seq_no": r[cols["seq_no"]],
            "ministry_name": r[cols["ministry_name"]],
            "ministry_code": r[ministry_code_col],
        }
        for group in ("recurring", "capital", "total"):
            for measure, idx in cols[group].items():
                record[f"{group}_{measure}"] = r[idx]
        records.append(record)
    return records


def extract_ocsc(raw_file: Path, config: dict) -> list[dict]:
    rows = load_sheet_rows(raw_file, config["sheet_name"])
    cols = config["columns"]

    def make_record(row, hierarchy_level, parent_category):
        return {
            "category_name": row[cols["category_name"]],
            "headcount": row[cols["headcount"]],
            "percent": row[cols["percent"]],
            "hierarchy_level": hierarchy_level,
            "parent_category": parent_category,
        }

    records = []

    grand_row = rows[config["grand_total_row"]]
    grand_name = grand_row[cols["category_name"]]
    records.append(make_record(grand_row, 0, None))

    for subtotal_row_idx, (leaf_start, leaf_end) in zip(
        config["subtotal_rows"], config["leaf_row_ranges"]
    ):
        subtotal_row = rows[subtotal_row_idx]
        subtotal_name = subtotal_row[cols["category_name"]]
        records.append(make_record(subtotal_row, 1, grand_name))

        for row in rows[leaf_start: leaf_end + 1]:
            records.append(make_record(row, 2, subtotal_name))

    return records


def write_csv(records: list[dict], out_path: Path) -> None:
    STAGING_DIR.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(records[0].keys()))
        writer.writeheader()
        writer.writerows(records)


def run_cgd() -> Path:
    config = load_config("cgd")
    raw_file = find_raw_file("cgd")
    # Step 7 (ข้อ 3c): a newly downloaded file's structure is never assumed
    # to match config/cgd.yaml — this raises loudly (StructureValidationError)
    # if it doesn't, instead of silently parsing wrong columns.
    validate_cgd_structure(raw_file, config)
    records = extract_cgd(raw_file, config)
    out_path = STAGING_DIR / "cgd_disbursement.csv"
    write_csv(records, out_path)
    print(f"{out_path.name}: {len(records)} rows")
    return out_path


def run_ocsc() -> Path:
    config = load_config("ocsc")
    raw_file = find_raw_file("ocsc")
    records = extract_ocsc(raw_file, config)
    out_path = STAGING_DIR / "ocsc_workforce.csv"
    write_csv(records, out_path)
    print(f"{out_path.name}: {len(records)} rows")
    return out_path


def main() -> None:
    run_cgd()
    run_ocsc()


if __name__ == "__main__":
    main()
