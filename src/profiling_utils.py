"""
Shared helper functions for EDA / data profiling scripts.

Used by eda_cgd.py and eda_ocsc.py. Kept intentionally small: just the
handful of primitives both scripts need when reading messy government
report-style Excel files (title rows, multi-row headers, footnotes,
possible live formulas).
"""

from openpyxl import load_workbook


def load_sheet_rows(path, sheet_name, data_only=False):
    """
    Load one sheet as a list of row-tuples. No pandas, no header guessing —
    we want to see the raw shape before deciding how to parse it.

    data_only=False -> formula cells return the formula text, e.g. '=F6*100/B6'
    data_only=True  -> formula cells return Excel's last CACHED numeric result

    We use data_only=False specifically to DETECT which cells are formulas
    (that's a data-quality issue worth flagging), not to read values.
    """
    wb = load_workbook(path, read_only=True, data_only=data_only)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def trim_trailing_none(row):
    """Drop trailing None cells so a printed row isn't cluttered with empty columns."""
    vals = list(row)
    while vals and vals[-1] is None:
        vals.pop()
    return vals


def count_formula_cells(path, sheet_name):
    """
    Count how many non-empty cells in a sheet are live Excel formulas
    (text starting with '='). Returns (formula_count, non_empty_count).
    """
    rows = load_sheet_rows(path, sheet_name, data_only=False)
    formula_count = 0
    non_empty = 0
    for row in rows:
        for cell in row:
            if cell is None:
                continue
            non_empty += 1
            if isinstance(cell, str) and cell.startswith('='):
                formula_count += 1
    return formula_count, non_empty


def classify_sheet_density(path, sheet_name, sample_rows=500):
    """
    Quick classification of one sheet: how many rows actually contain data
    vs. being empty (chart-only sheet) or just a title line.

    Returns a dict: {sheet, rows_with_data, first_cell_text}
    """
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet_name]
    rows_with_data = 0
    first_cell_text = ''
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= sample_rows:
            break
        vals = [v for v in row if v is not None]
        if vals:
            rows_with_data += 1
            if not first_cell_text:
                first_cell_text = str(vals[0])[:60]
    wb.close()
    return {
        'sheet': sheet_name,
        'rows_with_data': rows_with_data,
        'first_cell_text': first_cell_text,
    }
