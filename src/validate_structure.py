"""Step 7: schema/structure validation (ข้อ 3c). Runs before extract.py
trusts a file's layout -- raises loudly on any mismatch instead of
silently reading wrong columns."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


class StructureValidationError(Exception):
    """Raised when a workbook's layout doesn't match its config -- meant
    to stop the pipeline, not be downgraded to a warning."""


def validate_cgd_structure(file_path: Path, config: dict) -> None:
    """Raise StructureValidationError if file_path's sheet doesn't match config."""
    sheet_name = config["sheet_name"]

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise StructureValidationError(
                f"{file_path.name}: expected sheet {sheet_name!r} not found. "
                f"Sheets present: {wb.sheetnames}"
            )

        rows = list(wb[sheet_name].iter_rows(values_only=True))

        for anchor in config["validation"]["header_anchors"]:
            row_idx, col_idx, expected = anchor["row"], anchor["col"], anchor["text"]
            if row_idx >= len(rows):
                raise StructureValidationError(
                    f"{file_path.name}: sheet {sheet_name!r} has only {len(rows)} rows, "
                    f"expected a header row at index {row_idx}"
                )
            row = rows[row_idx]
            actual = row[col_idx] if col_idx < len(row) else None
            if actual != expected:
                raise StructureValidationError(
                    f"{file_path.name}: header mismatch at row {row_idx}, col {col_idx} "
                    f"— expected {expected!r}, found {actual!r}. The sheet's structure "
                    "may have changed; refusing to parse until config/cgd.yaml is "
                    "reviewed and updated to match."
                )

        total_row_idx = config["total_row"]
        if total_row_idx >= len(rows):
            raise StructureValidationError(
                f"{file_path.name}: expected a total row at index {total_row_idx}, but "
                f"sheet {sheet_name!r} only has {len(rows)} rows"
            )
        total_col = config["validation"]["total_row_label_col"]
        expected_total_label = config["validation"]["total_row_label_text"]
        total_row = rows[total_row_idx]
        actual_total_label = total_row[total_col] if total_col < len(total_row) else None
        if actual_total_label != expected_total_label:
            raise StructureValidationError(
                f"{file_path.name}: expected total-row label {expected_total_label!r} at "
                f"row {total_row_idx}, col {total_col}, found {actual_total_label!r} — "
                "the total row may have moved to a different position."
            )
    finally:
        wb.close()
