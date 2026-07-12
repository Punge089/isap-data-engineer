"""Step 7 downloader tests. Fixture-based, no live network calls (the one
real live download_cgd() run is documented in HANDOFF.md). Pure decision
logic (process_download, hash_exists_in_manifest, extract_download_url,
unwrap_response_bytes) is tested directly with no I/O.

fake_new_cgd_report.xlsx is plain placeholder bytes (not a zip) --
fake_new_cgd_report_zip_wrapped.xlsx mirrors the real ZIP-wrapped response
structure found live on 2026-07-12 (see unwrap_response_bytes tests)."""

import hashlib
import io
import json
import zipfile
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

import detect
import download as download_module

FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"

# Derived from the REAL fixture at test-collection time (not hand-transcribed
# from memory): loads tests/fixtures/cgd_listing_sample.html — the same real
# saved page from Step 6 — through detect.py's actual parser, so this is
# exactly the href BeautifulSoup produces from the genuine file (HTML
# entities like &amp; already decoded to &). If the fixture ever changes,
# this constant changes with it automatically instead of silently drifting
# out of sync with a copy-pasted string.
_REAL_FIXTURE_REPORT = detect.find_latest_cgd_report(
    (FIXTURES_DIR / "cgd_listing_sample.html").read_text(encoding="utf-8")
)
REAL_HREF = _REAL_FIXTURE_REPORT.download_href


def make_report(report_date: date, href: str = REAL_HREF) -> detect.CgdReport:
    return detect.CgdReport(title="ผลการเบิกจ่ายเงิน ณ ทดสอบ", report_date=report_date, download_href=href)


def make_template_entry() -> dict:
    return {
        "source": "CGD",
        "source_name_th": "กรมบัญชีกลาง - ผลการเบิกจ่ายงบประมาณ",
        "source_url": "https://www.cgd.go.th/x",
    }


# ---------------------------------------------------------------- pure logic

def test_real_fixture_href_is_the_javascript_opendownload_shape():
    """Sanity check on REAL_HREF itself: confirms it really did come out of
    the real fixture as a javascript:openDownload(...) call (the Step 6
    finding), not some other shape — if this ever fails, the fixture file
    changed in a way that invalidates every test below that relies on it."""
    assert REAL_HREF.startswith("javascript:openDownload(")
    assert "1428635636937" in REAL_HREF  # the real report's internal id, from the real page


def test_extract_download_url_from_real_fixture_href_exact_expected_url():
    """Standalone test against the exact real string extracted from the
    Step 6 fixture, with an explicit expected-URL assertion (not just
    'is not None')."""
    url = download_module.extract_download_url(REAL_HREF)
    assert url == (
        "https://www.cgd.go.th/cs/Satellite?blobcol=urldata&blobkey=id&"
        "blobtable=MungoBlobs&blobwhere=1438191564296&ssbinary=true"
    )


def test_extract_download_url_returns_none_on_unrecognized_href():
    assert download_module.extract_download_url("https://example.com/plain-link.xlsx") is None
    assert download_module.extract_download_url("javascript:somethingElse();") is None


# ---------------------------------------------------------------- zip-wrapped response
# Found for real against the live CGD site on 2026-07-12: the download link
# sometimes serves a ZIP containing a dated subfolder ('2026.07.03/...')
# with the actual .xlsx inside, not a bare .xlsx. tests/fixtures/
# fake_new_cgd_report_zip_wrapped.xlsx mirrors that exact structure with a
# genuinely valid, openable inner workbook (unlike fake_new_cgd_report.xlsx,
# which is plain placeholder bytes and would never have caught this bug).

def test_unwrap_response_bytes_extracts_xlsx_from_zip_wrapper():
    zip_bytes = (FIXTURES_DIR / "fake_new_cgd_report_zip_wrapped.xlsx").read_bytes()

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        expected_inner_bytes = zf.read("2026.07.03/2026.07.03.xlsx")

    result = download_module.unwrap_response_bytes(zip_bytes)

    assert result == expected_inner_bytes
    # the unwrapped bytes must be a real, openable workbook, not just "some bytes"
    wb = load_workbook(io.BytesIO(result))
    assert wb.sheetnames == ["2.กระทรวง"]


def test_unwrap_response_bytes_passes_plain_xlsx_through_unchanged():
    plain_bytes = (FIXTURES_DIR / "fake_new_cgd_report.xlsx").read_bytes()
    assert zipfile.is_zipfile(io.BytesIO(plain_bytes)) is False  # sanity: fixture is NOT a zip

    result = download_module.unwrap_response_bytes(plain_bytes)

    assert result == plain_bytes


def test_unwrap_response_bytes_raises_on_zip_with_no_xlsx_entry():
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("readme.txt", "no xlsx in here")
    zip_bytes_without_xlsx = buf.getvalue()

    with pytest.raises(ValueError, match="no .xlsx entry"):
        download_module.unwrap_response_bytes(zip_bytes_without_xlsx)


def test_hash_exists_in_manifest_true_and_false():
    manifest = {"files": [{"source": "CGD", "sha256": "abc123"}]}
    assert download_module.hash_exists_in_manifest("abc123", manifest) is True
    assert download_module.hash_exists_in_manifest("different", manifest) is False


def test_process_download_duplicate_hash_is_skipped():
    """A file whose content hash already exists in the manifest must be
    skipped -- not re-saved, not re-added -- even if detect.py thought the
    date was new (e.g. the same report re-published under a new date)."""
    file_bytes = b"identical content"
    existing_hash = hashlib.sha256(file_bytes).hexdigest()
    manifest = {"files": [{"source": "CGD", "report_date": "2026-06-05", "sha256": existing_hash}]}

    report = make_report(date(2026, 7, 3))
    result = download_module.process_download(report, file_bytes, manifest, make_template_entry())

    assert result is None


def test_process_download_new_hash_produces_new_entry_not_replacement():
    """A new hash must produce a brand-new entry dict — and the existing
    manifest data must be completely untouched, since process_download is
    pure decision logic; appending happens elsewhere."""
    old_bytes = b"old report content"
    old_hash = hashlib.sha256(old_bytes).hexdigest()
    manifest = {"files": [{"source": "CGD", "report_date": "2026-06-05", "sha256": old_hash}]}
    manifest_before = json.dumps(manifest, sort_keys=True)

    new_bytes = b"brand new report content"
    report = make_report(date(2026, 7, 3))
    result = download_module.process_download(report, new_bytes, manifest, make_template_entry())

    assert result is not None
    assert result["source"] == "CGD"
    assert result["report_date"] == "2026-07-03"
    assert result["local_path"] == "raw/cgd/2026_07_03.xlsx"
    assert result["sha256"] == hashlib.sha256(new_bytes).hexdigest()
    assert result["downloaded_by"] == "auto"
    # the manifest dict passed in was not mutated by process_download
    assert json.dumps(manifest, sort_keys=True) == manifest_before


# ---------------------------------------------------------------- manifest I/O

def test_append_manifest_entry_appends_not_replaces(tmp_path, monkeypatch):
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(
        json.dumps({"files": [{"source": "CGD", "sha256": "existing-hash"}]}), encoding="utf-8"
    )
    monkeypatch.setattr(download_module, "MANIFEST_PATH", manifest_path)

    new_entry = {"source": "CGD", "sha256": "new-hash"}
    download_module.append_manifest_entry(new_entry)

    updated = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert len(updated["files"]) == 2
    assert updated["files"][0]["sha256"] == "existing-hash"  # untouched
    assert updated["files"][1]["sha256"] == "new-hash"  # appended


# ---------------------------------------------------------------- OCSC refusal

def test_download_ocsc_refuses_without_any_network_call(monkeypatch):
    def _fail_if_called(*args, **kwargs):
        raise AssertionError("download_ocsc() must not make any network call")

    monkeypatch.setattr(download_module.requests, "get", _fail_if_called)
    monkeypatch.setattr(detect, "fetch_html", _fail_if_called)

    result = download_module.download_ocsc()

    assert "no automated download path exists" in result
    assert detect.OCSC_URL in result


# ---------------------------------------------------------------- full flow (fixture-based, not live)

def test_download_cgd_full_flow_with_fixture_bytes_not_live_network(tmp_path, monkeypatch):
    """Proves the download+manifest-append path end-to-end WITHOUT hitting
    the live network: the listing page and the file download are both
    faked, using tests/fixtures/fake_new_cgd_report.xlsx's bytes to stand
    in for a real download. This is stated explicitly, not hidden: there
    was nothing new on the real site to trigger this path for real (see
    module docstring and HANDOFF.md).
    """
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "files": [
                    {
                        "source": "CGD",
                        "source_name_th": "กรมบัญชีกลาง - ผลการเบิกจ่ายงบประมาณ",
                        "source_url": "https://www.cgd.go.th/x",
                        "local_path": "raw/cgd/2026_07_03.xlsx",
                        "report_date": "2026-07-03",
                        "sha256": "old-existing-hash",
                        "downloaded_at": "2026-07-09T00:00:00+07:00",
                        "downloaded_by": "manual",
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(download_module, "MANIFEST_PATH", manifest_path)
    monkeypatch.setattr(download_module, "REPO_ROOT", tmp_path)

    fake_listing_html = """
    <table><tr>
      <td>img</td>
      <td><h2 class="news-title"><a href="javascript:openDownload('NewsStat_C','999','/cs/Satellite?blobwhere=999');">ผลการเบิกจ่ายเงิน ณ วันที่ 10 สิงหาคม 2569</a></h2></td>
      <td>10/08/2569</td>
    </tr></table>
    """
    monkeypatch.setattr(detect, "fetch_html", lambda url: fake_listing_html)

    fixture_bytes = (FIXTURES_DIR / "fake_new_cgd_report.xlsx").read_bytes()

    class FakeResponse:
        content = fixture_bytes

        def raise_for_status(self):
            pass

    monkeypatch.setattr(download_module.requests, "get", lambda *a, **k: FakeResponse())

    result = download_module.download_cgd()

    assert "downloaded new file" in result
    assert "2026-08-10" in result

    saved_path = tmp_path / "raw" / "cgd" / "2026_08_10.xlsx"
    assert saved_path.exists()
    assert saved_path.read_bytes() == fixture_bytes

    updated_manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert len(updated_manifest["files"]) == 2  # appended, not replaced
    assert updated_manifest["files"][0]["sha256"] == "old-existing-hash"  # untouched
    new_entry = updated_manifest["files"][1]
    assert new_entry["report_date"] == "2026-08-10"
    assert new_entry["local_path"] == "raw/cgd/2026_08_10.xlsx"
    assert new_entry["sha256"] == hashlib.sha256(fixture_bytes).hexdigest()
    assert new_entry["downloaded_by"] == "auto"


def test_download_cgd_full_flow_unwraps_zip_wrapped_response(tmp_path, monkeypatch):
    """Same end-to-end shape as the plain-xlsx full-flow test above, but the
    fake network response is ZIP-wrapped like the real live CGD response
    found on 2026-07-12 -- proves download_cgd() unwraps it before hashing/
    saving/appending, and that the saved file is a valid, openable xlsx
    (not the outer zip bytes)."""
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "files": [
                    {
                        "source": "CGD",
                        "source_name_th": "กรมบัญชีกลาง - ผลการเบิกจ่ายงบประมาณ",
                        "source_url": "https://www.cgd.go.th/x",
                        "local_path": "raw/cgd/2026_07_03.xlsx",
                        "report_date": "2026-07-03",
                        "sha256": "old-existing-hash",
                        "downloaded_at": "2026-07-09T00:00:00+07:00",
                        "downloaded_by": "manual",
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(download_module, "MANIFEST_PATH", manifest_path)
    monkeypatch.setattr(download_module, "REPO_ROOT", tmp_path)

    fake_listing_html = """
    <table><tr>
      <td>img</td>
      <td><h2 class="news-title"><a href="javascript:openDownload('NewsStat_C','999','/cs/Satellite?blobwhere=999');">ผลการเบิกจ่ายเงิน ณ วันที่ 10 สิงหาคม 2569</a></h2></td>
      <td>10/08/2569</td>
    </tr></table>
    """
    monkeypatch.setattr(detect, "fetch_html", lambda url: fake_listing_html)

    zip_bytes = (FIXTURES_DIR / "fake_new_cgd_report_zip_wrapped.xlsx").read_bytes()
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        expected_inner_bytes = zf.read("2026.07.03/2026.07.03.xlsx")

    class FakeResponse:
        content = zip_bytes

        def raise_for_status(self):
            pass

    monkeypatch.setattr(download_module.requests, "get", lambda *a, **k: FakeResponse())

    result = download_module.download_cgd()

    assert "downloaded new file" in result
    assert "2026-08-10" in result

    saved_path = tmp_path / "raw" / "cgd" / "2026_08_10.xlsx"
    assert saved_path.exists()
    # saved bytes are the UNWRAPPED inner xlsx, not the outer zip container
    assert saved_path.read_bytes() == expected_inner_bytes
    assert zipfile.is_zipfile(saved_path)  # still true: a real xlsx IS a zip internally
    wb = load_workbook(saved_path)  # must not raise -- proves it's a valid, openable workbook
    assert wb.sheetnames == ["2.กระทรวง"]

    updated_manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert len(updated_manifest["files"]) == 2  # appended, not replaced
    new_entry = updated_manifest["files"][1]
    assert new_entry["report_date"] == "2026-08-10"
    assert new_entry["sha256"] == hashlib.sha256(expected_inner_bytes).hexdigest()  # hash of INNER bytes


def test_download_cgd_reports_nothing_new_without_downloading_file(tmp_path, monkeypatch):
    """If the fake listing page's latest date matches the manifest exactly,
    download_cgd() must stop before ever attempting the file download."""
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "files": [
                    {
                        "source": "CGD",
                        "source_name_th": "x",
                        "source_url": "https://www.cgd.go.th/x",
                        "local_path": "raw/cgd/2026_07_03.xlsx",
                        "report_date": "2026-07-03",
                        "sha256": "existing-hash",
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(download_module, "MANIFEST_PATH", manifest_path)

    fake_listing_html = """
    <table><tr>
      <td>img</td>
      <td><h2 class="news-title"><a href="javascript:openDownload('a','b','/cs/Satellite?x');">ผลการเบิกจ่ายเงิน ณ วันที่ 3 กรกฎาคม 2569</a></h2></td>
      <td>03/07/2569</td>
    </tr></table>
    """
    monkeypatch.setattr(detect, "fetch_html", lambda url: fake_listing_html)

    def _fail_if_called(*args, **kwargs):
        raise AssertionError("must not attempt the file download when nothing is new")

    monkeypatch.setattr(download_module.requests, "get", _fail_if_called)

    result = download_module.download_cgd()
    assert "nothing new" in result
