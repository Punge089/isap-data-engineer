"""Step 7 structure-validation tests.

The real-file test is the happy path (sanity check that today's committed
raw file still validates). The mangled-sheet tests are the actual point of
this module: they build a deliberately wrong fake .xlsx on the fly with
openpyxl and confirm validate_cgd_structure raises StructureValidationError
instead of silently accepting it — proving the negative path, not just the
positive one.
"""

from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook

from validate_structure import StructureValidationError, validate_cgd_structure

REPO_ROOT = Path(__file__).resolve().parent.parent

# The real header rows (3 and 4), exactly as they appear in the real file —
# used to build fake sheets that are correct EXCEPT for the one thing each
# test deliberately breaks.
REAL_ROW_3 = [
    "ลำดับ\nที่", "กระทรวง", None, "รายจ่ายประจำ", None, None, None, None, None,
    "รายจ่ายลงทุน", None, None, None, None, None, "รวม", None, None, None, None, None,
    "รหัสกระทรวง",
]
REAL_ROW_4 = [
    None, None, None, "วงเงิน\nงบประมาณ\nหลังโอน\nเปลี่ยนแปลง", "จัดสรร",
    "แผนการ\nใช้จ่าย", "PO+สำรอง\nเงินมีหนี้", "เบิกจ่าย",
    "ร้อยละ\nเบิกจ่าย\nต่องบ\nประมาณ\nหลังโอน\nเปลี่ยนแปลง",
]


def real_cgd_config() -> dict:
    return yaml.safe_load((REPO_ROOT / "config" / "cgd.yaml").read_text(encoding="utf-8"))


def build_fake_workbook(sheet_name: str, row_3, row_4, total_row_idx: int, total_row) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for _ in range(3):
        ws.append([None])  # rows 0-2: title block, irrelevant to validation
    ws.append(row_3)
    ws.append(row_4)
    while ws.max_row < total_row_idx:
        ws.append([None])
    ws.append(total_row)
    return wb


def test_validate_cgd_structure_passes_on_real_committed_file():
    config = real_cgd_config()
    raw_file = sorted((REPO_ROOT / "raw" / "cgd").glob("*.xlsx"))[-1]
    validate_cgd_structure(raw_file, config)  # must not raise


def test_validate_cgd_structure_raises_on_missing_sheet(tmp_path):
    config = real_cgd_config()
    wb = Workbook()
    wb.active.title = "not the right sheet name"
    fake_path = tmp_path / "wrong_sheet.xlsx"
    wb.save(fake_path)

    with pytest.raises(StructureValidationError, match="not found"):
        validate_cgd_structure(fake_path, config)


def test_validate_cgd_structure_raises_on_mangled_header(tmp_path):
    """Deliberately reshaped header: column 3 says something else instead
    of 'รายจ่ายประจำ' — simulating a real layout change."""
    config = real_cgd_config()
    mangled_row_3 = list(REAL_ROW_3)
    mangled_row_3[3] = "หมวดใหม่ (เปลี่ยนชื่อ)"  # wrong label

    wb = build_fake_workbook(
        config["sheet_name"], mangled_row_3, REAL_ROW_4,
        total_row_idx=29, total_row=["รวม"],
    )
    fake_path = tmp_path / "mangled_header.xlsx"
    wb.save(fake_path)

    with pytest.raises(StructureValidationError, match="header mismatch"):
        validate_cgd_structure(fake_path, config)


def test_validate_cgd_structure_raises_on_missing_column_entirely(tmp_path):
    """Deliberately shorter header row — as if a whole column got deleted,
    shifting everything after it (config expects col 21 = 'รหัสกระทรวง',
    but the row only has 10 cells now)."""
    config = real_cgd_config()
    truncated_row_3 = REAL_ROW_3[:10]

    wb = build_fake_workbook(
        config["sheet_name"], truncated_row_3, REAL_ROW_4,
        total_row_idx=29, total_row=["รวม"],
    )
    fake_path = tmp_path / "truncated_header.xlsx"
    wb.save(fake_path)

    with pytest.raises(StructureValidationError, match="header mismatch"):
        validate_cgd_structure(fake_path, config)


def test_validate_cgd_structure_raises_on_moved_total_row(tmp_path):
    """Header is correct, but the total row's label is missing from where
    config says it should be — simulating the total row shifting position
    (e.g. an extra row inserted above it)."""
    config = real_cgd_config()

    wb = build_fake_workbook(
        config["sheet_name"], REAL_ROW_3, REAL_ROW_4,
        total_row_idx=29, total_row=["ไม่ใช่แถวรวม"],  # wrong label, not 'รวม'
    )
    fake_path = tmp_path / "moved_total_row.xlsx"
    wb.save(fake_path)

    with pytest.raises(StructureValidationError, match="total-row label"):
        validate_cgd_structure(fake_path, config)


def test_validate_cgd_structure_raises_on_sheet_too_short_for_total_row(tmp_path):
    """Sheet ends before it ever reaches the row index the config expects
    the total row at — a different failure mode than a wrong label."""
    config = real_cgd_config()

    wb = Workbook()
    ws = wb.active
    ws.title = config["sheet_name"]
    for _ in range(3):
        ws.append([None])
    ws.append(REAL_ROW_3)
    ws.append(REAL_ROW_4)
    # sheet stops here -- nowhere near row index 29
    fake_path = tmp_path / "too_short.xlsx"
    wb.save(fake_path)

    with pytest.raises(StructureValidationError, match="only has"):
        validate_cgd_structure(fake_path, config)
